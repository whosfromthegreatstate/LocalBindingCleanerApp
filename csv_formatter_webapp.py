import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

st.set_page_config(page_title="Local Binding Formatter", layout="centered")
st.title("📄  Local Binding Asana to Production Formatter")
st.markdown("Upload your `.csv` file below. The app will clean the data and show a preview before download.")

uploaded_file = st.file_uploader("Upload CSV", type=["csv"])

if uploaded_file:
    # Read CSV
    df = pd.read_csv(uploaded_file)

    # Remove Column A (first column)
    df = df.iloc[:, 1:]

    # Fill down the Section/Column
    if 'Section/Column' in df.columns:
        df['Section/Column'] = df['Section/Column'].fillna(method='ffill')

    # Split Name into Name + Quantity
    def split_name_quantity(name):
        if pd.isna(name):
            return pd.Series([None, None])
        parts = str(name).split()
        name_parts = [p for p in parts if not p.isdigit()]
        quantity_parts = [p for p in parts if p.isdigit()]
        cleaned_name = " ".join(name_parts).strip() if name_parts else None
        quantity = int(quantity_parts[0]) if quantity_parts else None
        return pd.Series([cleaned_name, quantity])

    if 'Name' in df.columns:
        df[['Name', 'Quantity']] = df['Name'].apply(split_name_quantity)

    # Reorder columns
    cols = df.columns.tolist()
    if 'Quantity' in cols:
        cols.insert(cols.index('Name') + 1, cols.pop(cols.index('Quantity')))
        df = df[cols]

    # Fixed parent-child inheritance logic
    if {'Parent task', 'Name', 'Tags', 'Notes'}.issubset(df.columns):
        # Clean up the data first
        df['Tags'] = df['Tags'].fillna('').astype(str)
        df['Notes'] = df['Notes'].fillna('').astype(str)
        df['Parent task'] = df['Parent task'].fillna('').astype(str)
        
        # Create a simple mapping from Name to Tags and Notes
        name_to_tags = dict(zip(df['Name'], df['Tags']))
        name_to_notes = dict(zip(df['Name'], df['Notes']))
        
        # Apply inheritance to each child row
        for idx, row in df.iterrows():
            if row['Parent task'] and row['Parent task'].strip():  # This is a child row
                parent_name = row['Parent task'].strip()
                
                # Get parent's tags and notes
                parent_tags = name_to_tags.get(parent_name, '')
                parent_notes = name_to_notes.get(parent_name, '')
                
                # Get child's existing values
                child_tags = row['Tags'] if row['Tags'] and row['Tags'].strip() else ''
                child_notes = row['Notes'] if row['Notes'] and row['Notes'].strip() else ''
                
                # Combine tags (child first, then parent)
                combined_tags = []
                if child_tags:
                    combined_tags.extend([tag.strip() for tag in child_tags.split(',') if tag.strip()])
                if parent_tags:
                    parent_tag_list = [tag.strip() for tag in parent_tags.split(',') if tag.strip()]
                    for tag in parent_tag_list:
                        if tag not in combined_tags:  # Avoid duplicates
                            combined_tags.append(tag)
                
                # Combine notes (child first, then parent)
                combined_notes = []
                if child_notes:
                    combined_notes.append(child_notes)
                if parent_notes:
                    combined_notes.append(parent_notes)
                
                # Update the dataframe
                df.at[idx, 'Tags'] = ', '.join(combined_tags) if combined_tags else ''
                df.at[idx, 'Notes'] = '\n'.join(combined_notes) if combined_notes else ''

    # Preview cleaned data
    st.subheader("🔍 Preview of Cleaned Data")
    st.dataframe(df, use_container_width=True)

    # Prepare Excel with multiple sheets and formatting
    xlsx_output = BytesIO()
    with pd.ExcelWriter(xlsx_output, engine="openpyxl") as writer:
        def autofit_columns(worksheet):
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2

        def style_headers(worksheet):
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

        def apply_table_filter(worksheet):
            tab = Table(displayName="FilteredTable", ref=worksheet.dimensions)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            worksheet.add_table(tab)

        def hide_columns(worksheet, columns_to_hide):
            for col_letter in columns_to_hide:
                worksheet.column_dimensions[col_letter].hidden = True

        # Sheet 1: Full cleaned data
        df.to_excel(writer, index=False, sheet_name="Formatted Data")
        ws1 = writer.book["Formatted Data"]
        autofit_columns(ws1)
        style_headers(ws1)

        # Apply conditional formatting to rows with "Local Binding Shop Orders" in Projects column
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
            projects_cell = row[df.columns.get_loc('Projects')] if 'Projects' in df.columns else None
            name_cell = row[df.columns.get_loc('Name')] if 'Name' in df.columns else None
            if projects_cell and projects_cell.value == "Local Binding Shop Orders" and name_cell:
                name_cell.font = Font(bold=True, size=14, color="FFFFFF")
                name_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        # Sheet 2: Filtered & sorted copy
        filtered_df = df[df['Projects'].isna()] if 'Projects' in df.columns else df.copy()
        filtered_df = filtered_df.sort_values(by='Name', ascending=False, key=lambda col: col.str.lower())
        filtered_df.to_excel(writer, index=False, sheet_name="Filtered View")
        ws2 = writer.book["Filtered View"]
        autofit_columns(ws2)
        style_headers(ws2)
        apply_table_filter(ws2)
        hide_columns(ws2, ['A','B','G','H','I','J','O','P'])

        # Sheet 3: Pivot summary of sizes (if present)
        size_counts = pd.DataFrame()
        if 'Name' in df.columns:
            keywords = ['small', 'medium', 'large']
            for key in keywords:
                count = df[df['Name'].str.lower().str.contains(key, na=False)]['Quantity'].sum()
                size_counts.at[0, key.title()] = count if pd.notna(count) else 0
            size_counts = size_counts.T.reset_index()
            size_counts.columns = ['Size', 'Total Quantity']
            size_counts.to_excel(writer, index=False, sheet_name="Pivot Summary")
            ws3 = writer.book["Pivot Summary"]
            autofit_columns(ws3)
            style_headers(ws3)

        # Sheet 4: Detailed Analysis with filtering capabilities
        if all(col in df.columns for col in ['Name', 'Section/Column', 'Tags', 'Quantity']):
            # Create detailed analysis dataframe
            analysis_data = []
            
            for _, row in df.iterrows():
                if pd.notna(row['Quantity']) and row['Quantity'] > 0:
                    # Extract size from name
                    name_lower = str(row['Name']).lower()
                    size = 'Unknown'
                    if 'small' in name_lower or 'sm' in name_lower:
                        size = 'Small'
                    elif 'medium' in name_lower or 'med' in name_lower:
                        size = 'Medium'
                    elif 'large' in name_lower or 'lrg' in name_lower:
                        size = 'Large'
                    
                    # Extract color/pad type from tags
                    tags = str(row['Tags']).lower() if pd.notna(row['Tags']) else ''
                    color = 'No Color Specified'
                    
                    if 'purple' in tags:
                        color = 'Purple Pads'
                    elif 'black' in tags:
                        color = 'Black Pads'
                    elif 'blue' in tags or 'cerulean' in tags:
                        color = 'Blue Pads'
                    elif 'red' in tags or 'candy red' in tags:
                        color = 'Red Pads'
                    elif 'white' in tags or 'snow' in tags:
                        color = 'White Pads'
                    elif 'green' in tags:
                        color = 'Green Pads'
                    
                    analysis_data.append({
                        'Section/Column': row['Section/Column'] if pd.notna(row['Section/Column']) else 'No Section',
                        'Parent Name': row['Parent task'] if pd.notna(row['Parent task']) and row['Parent task'] else row['Name'],
                        'Item Name': row['Name'],
                        'Size': size,
                        'Color/Pad Type': color,
                        'Quantity': row['Quantity'],
                        'Tags': row['Tags'] if pd.notna(row['Tags']) else '',
                        'Notes': row['Notes'] if pd.notna(row['Notes']) else ''
                    })
            
            if analysis_data:
                analysis_df = pd.DataFrame(analysis_data)
                
                # Create summary pivot tables
                summary_data = []
                
                # Overall size summary
                size_summary = analysis_df.groupby('Size')['Quantity'].sum().reset_index()
                size_summary['Category'] = 'Overall Total'
                size_summary['Subcategory'] = size_summary['Size']
                summary_data.append(size_summary[['Category', 'Subcategory', 'Quantity']])
                
                # Size by color summary
                size_color_summary = analysis_df.groupby(['Size', 'Color/Pad Type'])['Quantity'].sum().reset_index()
                size_color_summary['Category'] = 'Size by Color'
                size_color_summary['Subcategory'] = size_color_summary['Size'] + ' - ' + size_color_summary['Color/Pad Type']
                summary_data.append(size_color_summary[['Category', 'Subcategory', 'Quantity']])
                
                # Section summary
                section_summary = analysis_df.groupby(['Section/Column', 'Size'])['Quantity'].sum().reset_index()
                section_summary['Category'] = 'By Section'
                section_summary['Subcategory'] = section_summary['Section/Column'] + ' - ' + section_summary['Size']
                summary_data.append(section_summary[['Category', 'Subcategory', 'Quantity']])
                
                # Combine all summaries
                final_summary = pd.concat(summary_data, ignore_index=True)
                
                # Write detailed analysis sheet
                analysis_df.to_excel(writer, index=False, sheet_name="Detailed Analysis")
                ws4 = writer.book["Detailed Analysis"]
                autofit_columns(ws4)
                style_headers(ws4)
                apply_table_filter(ws4)
                
                # Write summary sheet
                final_summary.to_excel(writer, index=False, sheet_name="Filterable Summary")
                ws5 = writer.book["Filterable Summary"]
                autofit_columns(ws5)
                style_headers(ws5)
                apply_table_filter(ws5)

    st.download_button(
        label="📥 Download Excel (.xlsx)",
        data=xlsx_output.getvalue(),
        file_name="cleaned_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )